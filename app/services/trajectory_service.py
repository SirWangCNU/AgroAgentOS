"""轨迹业务服务.

职责:
  - 轨迹 Excel 文件解析
  - 轨迹 CRUD 操作
  - Redis 缓存管理（轨迹列表、轨迹点、统计结果）

缓存策略:
  - 轨迹列表: 按 field_id 缓存，TTL 1小时
  - 轨迹点: 按 file_id 缓存，TTL 30分钟（数据量大，较短 TTL）
  - 统计结果: 按 file_id 缓存，TTL 2小时（计算密集，较长 TTL）
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Optional

from loguru import logger
from sqlalchemy.orm import Session

from app.core.redis import redis_manager
from app.core.sqlite import sqlite_manager
from app.exceptions import AppException
from app.models.farm import Field
from app.models.trajectory import TrajectoryFile, TrajectoryPoint
from app.schemas.trajectory import (
    TrajectoryFileInfo,
    TrajectoryPointsResponse,
    TrajectoryPointData,
    TrajectoryStatsResponse,
    TrajectoryUploadResponse,
)
from app.services import trajectory_analysis

# 缓存 TTL（秒）
TRAJECTORY_LIST_TTL = 3600      # 1小时
TRAJECTORY_POINTS_TTL = 1800    # 30分钟
TRAJECTORY_STATS_TTL = 7200     # 2小时


# ==================== Excel 解析 ====================

# 列名映射（支持中英文）
COLUMN_MAPPING = {
    "gps_time": ["gps_time", "GPS时间", "时间", "time", "datetime", "记录时间"],
    "latitude": ["latitude", "纬度", "lat", "y", "lat_wgs84", "lat_gcj02"],
    "longitude": ["longitude", "经度", "lng", "lon", "x", "lon_wgs84", "lon_gcj02"],
    "speed": ["speed", "速度", "velocity", "km/h", "速度(km/h)"],
    "work_status": ["work_status", "工作状态", "作业状态", "status", "state", "是否作业"],
    "depth": ["depth", "作业深度", "耕深", "深度", "depth_cm"],
    "depth_std": ["depth_std", "深度标准值", "标准深度", "目标深度"],
    "work_width": ["work_width", "幅宽", "width", "作业幅宽"],
    "machine_id": ["machine_id", "农机编号", "机器号", "tractor_id", "机具编号"],
}


def _is_valid_latlon(lat: float, lon: float) -> bool:
    """检查是否为有效的经纬度坐标（度为单位）."""
    # 经纬度范围检查
    if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
        return False
    # 排除 0,0 坐标（通常是无效数据）
    if lat == 0 and lon == 0:
        return False
    # 排除明显的投影坐标（值通常很大，如几十万）
    if abs(lat) > 90 or abs(lon) > 180:
        return False
    return True


def _gcj02_to_wgs84(lat: float, lon: float) -> tuple[float, float]:
    """GCJ-02 坐标转 WGS-84 坐标.

    中国国测局坐标系(GCJ-02)转 GPS 坐标系(WGS-84).
    """
    import math

    PI = math.pi
    a = 6378245.0  # 长半轴
    ee = 0.00669342162296594  # 偏心率平方

    def _transform_lat(x, y):
        ret = -100.0 + 2.0 * x + 3.0 * y + 0.2 * y * y + 0.1 * x * y + 0.2 * math.sqrt(abs(x))
        ret += (20.0 * math.sin(6.0 * x * PI) + 20.0 * math.sin(2.0 * x * PI)) * 2.0 / 3.0
        ret += (20.0 * math.sin(y * PI) + 40.0 * math.sin(y / 3.0 * PI)) * 2.0 / 3.0
        ret += (160.0 * math.sin(y / 12.0 * PI) + 320 * math.sin(y * PI / 30.0)) * 2.0 / 3.0
        return ret

    def _transform_lon(x, y):
        ret = 300.0 + x + 2.0 * y + 0.1 * x * x + 0.1 * x * y + 0.1 * math.sqrt(abs(x))
        ret += (20.0 * math.sin(6.0 * x * PI) + 20.0 * math.sin(2.0 * x * PI)) * 2.0 / 3.0
        ret += (20.0 * math.sin(x * PI) + 40.0 * math.sin(x / 3.0 * PI)) * 2.0 / 3.0
        ret += (150.0 * math.sin(x / 12.0 * PI) + 300.0 * math.sin(x / 30.0 * PI)) * 2.0 / 3.0
        return ret

    dlat = _transform_lat(lon - 105.0, lat - 35.0)
    dlon = _transform_lon(lon - 105.0, lat - 35.0)

    radlat = lat / 180.0 * PI
    magic = math.sin(radlat)
    magic = 1 - ee * magic * magic
    sqrtmagic = math.sqrt(magic)

    dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * PI)
    dlon = (dlon * 180.0) / (a / sqrtmagic * math.cos(radlat) * PI)

    return lat - dlat, lon - dlon


def _detect_columns(headers: list[str]) -> dict[str, int]:
    """自动检测列名映射.

    Args:
        headers: Excel 表头列表

    Returns:
        {字段名: 列索引} 映射
    """
    column_map = {}
    for field_name, aliases in COLUMN_MAPPING.items():
        for idx, header in enumerate(headers):
            if header and header.strip().lower() in [a.lower() for a in aliases]:
                column_map[field_name] = idx
                break
    return column_map


def _parse_work_status(value: Any) -> str:
    """解析工作状态，支持多种格式."""
    if not value:
        return "idle"
    val = str(value).strip().lower()
    # Working 状态
    if val in ("working", "作业", "工作中", "作业中", "1", "true", "是", "yes"):
        return "working"
    # Transporting 状态
    if val in ("transporting", "转移", "运输", "运输中"):
        return "transporting"
    # Idle 状态
    if val in ("idle", "空闲", "0", "false", "否", "no", "停止"):
        return "idle"
    return "idle"


def _parse_datetime(value: Any) -> Optional[datetime]:
    """解析日期时间."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        # 尝试多种格式
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"]:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def parse_excel(file_content: bytes, filename: str) -> dict[str, Any]:
    """解析 Excel 文件.

    Args:
        file_content: 文件内容（字节）
        filename: 文件名

    Returns:
        {
            "machine_id": str,
            "work_width": float,
            "points": [{"seq": int, "gps_time": datetime, "latitude": float, ...}, ...]
        }

    Raises:
        AppException: 解析失败
    """
    try:
        import openpyxl
        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active

        if not ws:
            raise AppException(status_code=400, detail="Excel 文件没有工作表")

        # 读取表头
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value) if cell.value else "")

        # 检测列映射
        col_map = _detect_columns(headers)

        # 必需列检查
        required = ["latitude", "longitude"]
        for field in required:
            if field not in col_map:
                raise AppException(
                    status_code=400,
                    detail=f"Excel 缺少必需列: {field} (支持的列名: {', '.join(COLUMN_MAPPING[field])})"
                )

        # 解析数据
        points = []
        machine_id = ""
        work_width = 0.0
        seq = 0

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳过空行
            if not row or all(cell is None for cell in row):
                continue

            # 提取经纬度
            try:
                lat_idx = col_map["latitude"]
                lon_idx = col_map["longitude"]
                latitude = float(row[lat_idx]) if row[lat_idx] is not None else None
                longitude = float(row[lon_idx]) if row[lon_idx] is not None else None
            except (ValueError, TypeError, IndexError):
                continue

            if latitude is None or longitude is None:
                continue

            # 检测是否为有效经纬度（度为单位）
            # 如果值很大，可能是投影坐标（米为单位），跳过
            if not _is_valid_latlon(latitude, longitude):
                # 尝试交换经纬度（有些数据源是 lon,lat 顺序）
                if _is_valid_latlon(longitude, latitude):
                    latitude, longitude = longitude, latitude
                    logger.debug(f"[Excel] 第 {row_idx} 行: 经纬度已交换")
                else:
                    logger.debug(f"[Excel] 第 {row_idx} 行: 坐标无效 ({latitude}, {longitude})，跳过")
                    continue

            seq += 1
            point = {
                "seq": seq,
                "latitude": latitude,
                "longitude": longitude,
            }

            # 可选字段
            if "gps_time" in col_map:
                point["gps_time"] = _parse_datetime(row[col_map["gps_time"]])
            if "speed" in col_map:
                try:
                    point["speed"] = float(row[col_map["speed"]]) if row[col_map["speed"]] else 0.0
                except (ValueError, TypeError):
                    point["speed"] = 0.0
            if "work_status" in col_map:
                point["work_status"] = _parse_work_status(row[col_map["work_status"]])
            if "depth" in col_map:
                try:
                    point["depth"] = float(row[col_map["depth"]]) if row[col_map["depth"]] else 0.0
                except (ValueError, TypeError):
                    point["depth"] = 0.0
            if "depth_std" in col_map:
                try:
                    point["depth_std"] = float(row[col_map["depth_std"]]) if row[col_map["depth_std"]] else 0.0
                except (ValueError, TypeError):
                    point["depth_std"] = 0.0

            points.append(point)

            # 提取机器号和幅宽（从第一行有效数据）
            if not machine_id and "machine_id" in col_map:
                machine_id = str(row[col_map["machine_id"]] or "")
            if not work_width and "work_width" in col_map:
                try:
                    work_width = float(row[col_map["work_width"]] or 0)
                except (ValueError, TypeError):
                    pass

        wb.close()

        if not points:
            raise AppException(status_code=400, detail="Excel 中没有有效的轨迹点数据")

        # 检测是否需要坐标转换（GCJ-02 -> WGS-84）
        # 如果列名包含 "gcj02"，则进行转换
        need_convert = False
        for col_name in ["latitude", "longitude"]:
            if col_name in col_map:
                original_header = headers[col_map[col_name]].lower()
                if "gcj02" in original_header or "gcj" in original_header:
                    need_convert = True
                    break

        if need_convert:
            logger.info(f"[Excel] 检测到 GCJ-02 坐标，转换为 WGS-84")
            for point in points:
                point["latitude"], point["longitude"] = _gcj02_to_wgs84(
                    point["latitude"], point["longitude"]
                )

        logger.info(f"[Excel] 解析完成: {filename}, {len(points)} 个轨迹点")
        return {
            "machine_id": machine_id,
            "work_width": work_width,
            "points": points,
        }

    except AppException:
        raise
    except Exception as e:
        logger.error(f"[Excel] 解析失败: {filename}, error={e}")
        raise AppException(status_code=400, detail=f"Excel 解析失败: {str(e)}")


# ==================== 轨迹 CRUD ====================


def upload_trajectory(
    field_id: int,
    user_id: int,
    file_content: bytes,
    filename: str,
    coord_system: str = "auto",
) -> TrajectoryUploadResponse:
    """上传并解析轨迹文件.

    Args:
        field_id: 地块 ID
        user_id: 用户 ID
        file_content: 文件内容
        filename: 文件名
        coord_system: 坐标系 (wgs84/gcj02/auto)

    Returns:
        TrajectoryUploadResponse
    """
    # 验证地块存在且属于当前用户
    _verify_field_access(field_id, user_id)

    # 解析 Excel
    parsed = parse_excel(file_content, filename)
    points_data = parsed["points"]

    # 根据 coord_system 参数处理坐标转换
    if coord_system == "gcj02":
        logger.info(f"[Excel] 强制 GCJ-02 -> WGS-84 转换")
        for point in points_data:
            point["latitude"], point["longitude"] = _gcj02_to_wgs84(
                point["latitude"], point["longitude"]
            )
    elif coord_system == "wgs84":
        # 不需要转换
        pass
    # auto 模式下，parse_excel 已经根据列名自动处理了

    # 计算统计数据
    stats = trajectory_analysis.calc_trajectory_stats(
        file_id=0,  # 临时，后面会更新
        filename=filename,
        machine_id=parsed["machine_id"],
        points=points_data,
        work_width=parsed["work_width"],
    )

    # 存入数据库
    with sqlite_manager.session() as sess:
        # 创建轨迹文件记录
        traj_file = TrajectoryFile(
            field_id=field_id,
            filename=filename,
            machine_id=parsed["machine_id"],
            point_count=len(points_data),
            start_time=stats.start_time,
            end_time=stats.end_time,
            total_distance_m=stats.total_distance_m,
            work_distance_m=stats.work_distance_m,
            work_area_mu=stats.work_area_mu,
            avg_depth=stats.avg_depth,
            avg_speed=stats.avg_speed,
            depth_std=stats.depth_std,
            work_width=parsed["work_width"],
        )
        sess.add(traj_file)
        sess.flush()
        file_id = traj_file.id

        # 批量插入轨迹点（分批，每批 500 条）
        batch_size = 500
        for i in range(0, len(points_data), batch_size):
            batch = points_data[i:i + batch_size]
            for p in batch:
                point = TrajectoryPoint(
                    file_id=file_id,
                    seq=p["seq"],
                    gps_time=p.get("gps_time"),
                    latitude=p["latitude"],
                    longitude=p["longitude"],
                    speed=p.get("speed", 0.0),
                    work_status=p.get("work_status", "idle"),
                    depth=p.get("depth", 0.0),
                    depth_std=p.get("depth_std", 0.0),
                )
                sess.add(point)
            sess.flush()

        # 更新统计（使用真实的 file_id）
        traj_file_info = TrajectoryFileInfo.model_validate(traj_file)
        sess.expunge(traj_file)

    # 清除该地块的轨迹列表缓存
    redis_manager.delete(redis_manager.trajectory_list_key(field_id))

    logger.info(f"[Trajectory] 上传成功: file_id={file_id}, field_id={field_id}, points={len(points_data)}")

    return TrajectoryUploadResponse(
        file_info=traj_file_info,
        message=f"轨迹上传成功，共解析 {len(points_data)} 个轨迹点"
    )


def get_trajectories(field_id: int, user_id: int) -> list[TrajectoryFileInfo]:
    """获取地块的轨迹列表（带缓存）.

    Args:
        field_id: 地块 ID
        user_id: 用户 ID

    Returns:
        轨迹文件信息列表
    """
    _verify_field_access(field_id, user_id)

    # 尝试从缓存获取
    cache_key = redis_manager.trajectory_list_key(field_id)
    cached = redis_manager.get(cache_key)
    if cached is not None:
        logger.debug(f"[Trajectory] 缓存命中: field_id={field_id}")
        return [TrajectoryFileInfo(**item) for item in cached]

    # 从数据库查询
    with sqlite_manager.session() as sess:
        files = (
            sess.query(TrajectoryFile)
            .filter(TrajectoryFile.field_id == field_id)
            .order_by(TrajectoryFile.created_at.desc())
            .all()
        )
        result = [TrajectoryFileInfo.model_validate(f) for f in files]

    # 写入缓存
    cache_data = [item.model_dump(mode="json") for item in result]
    redis_manager.set(cache_key, cache_data, expire=TRAJECTORY_LIST_TTL)

    return result


def get_trajectory_points(file_id: int, user_id: int) -> TrajectoryPointsResponse:
    """获取轨迹点数据（带缓存）.

    Args:
        file_id: 轨迹文件 ID
        user_id: 用户 ID

    Returns:
        TrajectoryPointsResponse
    """
    # 验证权限
    traj_file = _get_trajectory_file_with_auth(file_id, user_id)

    # 尝试从缓存获取
    cache_key = redis_manager.trajectory_points_key(file_id)
    cached = redis_manager.get(cache_key)
    if cached is not None:
        logger.debug(f"[Trajectory] 缓存命中: file_id={file_id} points")
        points = [TrajectoryPointData(**p) for p in cached]
        return TrajectoryPointsResponse(
            file_id=file_id,
            point_count=len(points),
            points=points,
        )

    # 从数据库查询
    with sqlite_manager.session() as sess:
        db_points = (
            sess.query(TrajectoryPoint)
            .filter(TrajectoryPoint.file_id == file_id)
            .order_by(TrajectoryPoint.seq)
            .all()
        )
        points = [TrajectoryPointData.model_validate(p) for p in db_points]

    # 写入缓存
    cache_data = [p.model_dump(mode="json") for p in points]
    redis_manager.set(cache_key, cache_data, expire=TRAJECTORY_POINTS_TTL)

    return TrajectoryPointsResponse(
        file_id=file_id,
        point_count=len(points),
        points=points,
    )


def get_trajectory_stats(file_id: int, user_id: int) -> TrajectoryStatsResponse:
    """获取轨迹统计分析（带缓存）.

    Args:
        file_id: 轨迹文件 ID
        user_id: 用户 ID

    Returns:
        TrajectoryStatsResponse
    """
    # 验证权限
    traj_file = _get_trajectory_file_with_auth(file_id, user_id)

    # 尝试从缓存获取
    cache_key = redis_manager.trajectory_stats_key(file_id)
    cached = redis_manager.get(cache_key)
    if cached is not None:
        logger.debug(f"[Trajectory] 缓存命中: file_id={file_id} stats")
        return TrajectoryStatsResponse(**cached)

    # 获取轨迹点并计算统计
    points_resp = get_trajectory_points(file_id, user_id)
    points_data = [p.model_dump(mode="json") for p in points_resp.points]

    stats = trajectory_analysis.calc_trajectory_stats(
        file_id=file_id,
        filename=traj_file.filename,
        machine_id=traj_file.machine_id,
        points=points_data,
        work_width=traj_file.work_width,
    )

    # 写入缓存
    redis_manager.set(cache_key, stats.model_dump(mode="json"), expire=TRAJECTORY_STATS_TTL)

    return stats


def delete_trajectory(file_id: int, user_id: int) -> None:
    """删除轨迹文件及其轨迹点.

    Args:
        file_id: 轨迹文件 ID
        user_id: 用户 ID
    """
    # 验证权限
    traj_file = _get_trajectory_file_with_auth(file_id, user_id)
    field_id = traj_file.field_id

    with sqlite_manager.session() as sess:
        # 删除轨迹点
        sess.query(TrajectoryPoint).filter(TrajectoryPoint.file_id == file_id).delete()
        # 删除轨迹文件
        sess.query(TrajectoryFile).filter(TrajectoryFile.id == file_id).delete()

    # 清除缓存
    redis_manager.delete(redis_manager.trajectory_list_key(field_id))
    redis_manager.delete(redis_manager.trajectory_points_key(file_id))
    redis_manager.delete(redis_manager.trajectory_stats_key(file_id))

    logger.info(f"[Trajectory] 删除成功: file_id={file_id}")


# ==================== 内部辅助 ====================


def _verify_field_access(field_id: int, user_id: int) -> Field:
    """验证地块存在且属于当前用户."""
    with sqlite_manager.session() as sess:
        field = sess.query(Field).filter(Field.id == field_id).first()
        if not field:
            raise AppException(status_code=404, detail="地块不存在")
        from app.models.farm import Farm
        farm = sess.query(Farm).filter(Farm.id == field.farm_id, Farm.user_id == user_id).first()
        if not farm:
            raise AppException(status_code=403, detail="无权访问该地块")
        return field


def _get_trajectory_file_with_auth(file_id: int, user_id: int) -> TrajectoryFile:
    """获取轨迹文件并验证权限."""
    with sqlite_manager.session() as sess:
        traj_file = sess.query(TrajectoryFile).filter(TrajectoryFile.id == file_id).first()
        if not traj_file:
            raise AppException(status_code=404, detail="轨迹文件不存在")
        # 验证所属地块的权限
        _verify_field_access(traj_file.field_id, user_id)
        return traj_file
